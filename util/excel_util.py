import openpyxl

def get_row_count(path, sheet_name):
    workbook = openpyxl.load_workbook(path, read_only=False)
    sheet = workbook[sheet_name]
    return sheet.max_row

def get_column_count(path, sheet_name):
    workbook = openpyxl.load_workbook(path, read_only=False)
    sheet = workbook[sheet_name]
    return sheet.max_column

def get_cell_data(path, sheet_name, row, col):
    workbook = openpyxl.load_workbook(path, read_only=False)
    sheet = workbook[sheet_name]
    return sheet.cell(row=row,column=col).value

def set_cell_data(path, sheet_name, row, col, data):
    workbook = openpyxl.load_workbook(path, read_only=False)
    sheet = workbook[sheet_name]
    sheet.cell(row=row,column=col).value = data
    workbook.save(path)
    
def get_excel_data(path, sheet_name):
    final_list = []
    workbook = openpyxl.load_workbook(path, read_only=False) 
    sheet = workbook[sheet_name] 
    total_rows = sheet.max_row
    total_cols = sheet.max_column

    for r in range(2, total_rows+1):
        row_list = []
        for c in range(1, total_cols+1):
            row_list.append(str(sheet.cell(row=r, column=c).value))
        final_list.append(row_list)

    return final_list 
    


